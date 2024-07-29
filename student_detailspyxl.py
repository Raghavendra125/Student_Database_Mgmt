
import openpyxl

# File path
excel_file = 'student01.xlsx'

def load_students():
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        students = []
        for row in sheet.iter_rows(values_only=True):
            students.append({
                "usn": row[0],
                "name": row[1],
                "branch": row[2],
                "year": row[3],
                "mobile_no":row[4]
            })
        workbook.close()
        return students
    except FileNotFoundError:
        # If file doesn't exist, return an empty list
        return []

def save_students(students):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["USN", "Name", "Branch", "Year","mobile_no"])
    for student in students:
        sheet.append([student["usn"], student["name"], student["branch"], student["year"],student["mobile_no"]])
    workbook.save(excel_file)

def add_student(usn, name, branch, year,mobile_no):
    students = load_students()
    students.append({"usn": usn, "name": name, "branch": branch, "year": year,"mobile_no":mobile_no})

    save_students(students)

def display_students():
    students = load_students()
    
    # Using a set to store unique student entries
    unique_students = set()
    
    print("Student Information:")
    for student in students:
        student_info = (student['usn'],"\t", student['name'],"\t", student['branch'],"\t", student['year'],"\t", student['mobile_no'])
        if student_info not in unique_students:
            unique_students.add(student_info)
            print(f"{student['usn']}\t{student['name']}\t\t{student['branch']}\t\t {student['year']}\t{student['mobile_no']}")


def update_student(usn, updated_name, updated_branch, updated_year,updated_mobile_no):
    students = load_students()
    for student in students:
        if student["usn"] == usn:
            student["name"] = updated_name
            student["branch"] = updated_branch
            student["year"] = updated_year
            student["mobile_no"]=updated_mobile_no
            save_students(students)
            print("Student updated successfully.")
            return
    print("Student not found.")

def delete_student(usn):
    students = load_students()
    for student in students:
        if student["usn"] == usn:
            students.remove(student)
            save_students(students)
            print("Student deleted successfully.")
            return
    print("Student not found.")

